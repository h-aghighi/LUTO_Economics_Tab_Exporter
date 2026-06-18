#!/usr/bin/env python3
r"""
Export LUTO2 Economics dashboard data to state-organised Excel workbooks.

The exporter reads the JavaScript data files created for the LUTO2 Economics
report and writes one workbook per state. Each workbook contains regional
worksheets and, where available, a state-total worksheet based on the direct
``region_state`` data.

The output includes the main Economics sections (Sum, Ag, Ag Mgt and Non-Ag),
profitability summaries, and renewable revenue/cost detail for Onshore Wind and
Utility Solar PV.

Source selection follows the dashboard configuration rather than values read
from rendered charts. In particular:

* Non-Ag cost uses ``Economics_overview_Non_Ag.js`` → ``Cost > ALL``.
* Ag Mgt cost uses the leaf series under ``ALL > ALL > ALL`` in
  ``Economics_Am_cost.js``.
* Onshore Wind and Utility Solar PV profit use the direct profit series in
  ``Economics_Am_profit.js``.
* The Sum profitability table uses ``Profit = Total revenue + Total cost``
  because costs are stored as negative values.
* Non-Ag revenue is set to zero only when the direct profit and cost values show
  that no separate revenue series is present; profitability is then left blank.

The script accepts either a single ``DATA_REPORT/data`` directory or a parent
folder containing multiple ``Run_*`` directories.

Examples
--------
Single run::

    python luto2_economics_state_excel_exporter.py ^
      --data-dir "C:\path\to\Run_G0001\DATA_REPORT\data" ^
      --output-dir "C:\path\to\outputs" ^
      --output-prefix Run_G0001 ^
      --states Queensland

Multiple runs::

    python luto2_economics_state_excel_exporter.py ^
      --reports-base-dir "C:\path\to\Report_Data" ^
      --output-dir "C:\path\to\outputs"

Requirements
------------
``pandas`` and ``openpyxl``
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Missing packages. Run: python -m pip install pandas openpyxl") from exc


# Cache parsed JavaScript objects because the same files are reused across regions.
_PARSED_JS_CACHE: Dict[Tuple[str, float], Tuple[Optional[str], Optional[Any], Optional[str]]] = {}


OVERVIEW_FILES: Dict[str, str] = {
    "Sum": "Economics_overview_sum.js",
    "Ag": "Economics_overview_Ag.js",
    "Ag Mgt": "Economics_overview_Am.js",
    "Non-Ag": "Economics_overview_Non_Ag.js",
}

# Direct group-level profit series used by the dashboard.
ECONOMICS_GROUP_PROFIT_FILE = "Economics_Sum.js"

GROUP_PROFIT_COLUMN = {
    "Ag": "Agricultural Land-use",
    "Ag Mgt": "Agricultural Management",
    "Non-Ag": "Non-Agricultural Land-use",
}

# Renewable columns appended to the Ag Mgt overview table.
AG_MGT_BREAKDOWN_TECHS = ["Onshore Wind", "Utility Solar PV"]
AG_MGT_DETAIL_COLUMNS = [
    ("Revenue", "Onshore Wind"),
    ("Revenue", "Utility Solar PV"),
    ("Costs (opex) ", "Onshore Wind"),
    ("Costs (opex) ", "Utility Solar PV"),
    ("Costs (capex) ", "Onshore Wind"),
    ("Costs (capex) ", "Utility Solar PV"),
    ("Costs (other) ", "Onshore Wind"),
    ("Costs (other) ", "Utility Solar PV"),
]

GROUP_ORDER = ["Sum", "Ag", "Ag Mgt", "Non-Ag"]

SERIES_ORDER_SUM = [
    "Agricultural Land-use (revenue)",
    "Agricultural Management (revenue)",
    "Non-Agricultural Land-use (revenue)",
    "Agricultural Land-use (cost)",
    "Agricultural Management (cost)",
    "Non-Agricultural Land-use (cost)",
    "Transition cost (Ag2Ag)",
    "Transition cost (Ag2Non-Ag)",
    "Profit",
]

REGION_TO_STATE = {
    # National / territories
    "AUSTRALIA": "National",
    "ACT": "Australian Capital Territory",
    "Northern Territory": "Northern Territory",

    # Queensland
    "Burdekin": "Queensland",
    "Burnett Mary": "Queensland",
    "Cape York": "Queensland",
    "Co-operative Management Area": "Queensland",
    "Condamine": "Queensland",
    "Desert Channels": "Queensland",
    "Fitzroy": "Queensland",
    "Mackay Whitsunday": "Queensland",
    "Maranoa Balonne and Border Rivers": "Queensland",
    "Northern Gulf": "Queensland",
    "South East Queensland": "Queensland",
    "South West Queensland": "Queensland",
    "Southern Gulf": "Queensland",
    "Torres Strait": "Queensland",
    "Wet Tropics": "Queensland",

    # New South Wales
    "Central Tablelands": "New South Wales",
    "Central West": "New South Wales",
    "Greater Sydney": "New South Wales",
    "Hunter": "New South Wales",
    "Murray": "New South Wales",
    "North Coast": "New South Wales",
    "North West NSW": "New South Wales",
    "Northern Tablelands": "New South Wales",
    "Riverina": "New South Wales",
    "South East NSW": "New South Wales",
    "Western": "New South Wales",

    # Victoria
    "Corangamite": "Victoria",
    "East Gippsland": "Victoria",
    "Glenelg Hopkins": "Victoria",
    "Goulburn Broken": "Victoria",
    "Mallee": "Victoria",
    "North Central": "Victoria",
    "North East": "Victoria",
    "Port Phillip and Western Port": "Victoria",
    "West Gippsland": "Victoria",
    "Wimmera": "Victoria",

    # South Australia
    "Adelaide and Mount Lofty Ranges": "South Australia",
    "Alinytjara Wilurara": "South Australia",
    "Eyre Peninsula": "South Australia",
    "Kangaroo Island": "South Australia",
    "Northern and Yorke": "South Australia",
    "South Australian Arid Lands": "South Australia",
    "South Australian Murray Darling Basin": "South Australia",
    "South East": "South Australia",

    # Western Australia
    "Avon River Basin": "Western Australia",
    "Northern Agricultural Region": "Western Australia",
    "Peel-Harvey Region": "Western Australia",
    "Rangelands Region": "Western Australia",
    "South Coast Region": "Western Australia",
    "South West Region": "Western Australia",
    "Swan Region": "Western Australia",

    # Tasmania
    "North NRM Region": "Tasmania",
    "North West NRM Region": "Tasmania",
    "South NRM Region": "Tasmania",
}


def log(message: str) -> None:
    print(message, flush=True)


def sanitize_filename(value: Any) -> str:
    text = re.sub(r'[*?:"<>|\\/]+', "_", str(value))
    text = re.sub(r"\s+", "_", text).strip("_ .")
    return text or "output"


def sanitize_sheet_name(value: Any) -> str:
    text = re.sub(r"[\[\]\*\?:/\\]", "_", str(value))
    text = re.sub(r"\s+", " ", text).strip(" .")
    return (text or "Sheet")[:31]


def unique_sheet_name(name: Any, used: set[str]) -> str:
    base = sanitize_sheet_name(name)
    candidate = base
    i = 1
    while candidate.lower() in used:
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(candidate.lower())
    return candidate


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def strip_comments_and_trailing_commas(text: str) -> str:
    out: List[str] = []
    i = 0
    in_str: Optional[str] = None
    escape = False

    while i < len(text):
        ch = text[i]
        nxt = text[i + 1] if i + 1 < len(text) else ""

        if in_str:
            out.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == in_str:
                in_str = None
            i += 1
            continue

        if ch in ("'", '"'):
            in_str = ch
            out.append(ch)
            i += 1
            continue

        if ch == "/" and nxt == "/":
            i += 2
            while i < len(text) and text[i] not in "\r\n":
                i += 1
            continue

        if ch == "/" and nxt == "*":
            i += 2
            while i + 1 < len(text) and not (text[i] == "*" and text[i + 1] == "/"):
                i += 1
            i += 2
            continue

        out.append(ch)
        i += 1

    cleaned = "".join(out)
    return re.sub(r",\s*([}\]])", r"\1", cleaned)


def js_like_to_json(text: str) -> str:
    text = strip_comments_and_trailing_commas(text)

    def replace_single(match: re.Match) -> str:
        inner = match.group(1).replace('"', '\\"')
        return f'"{inner}"'

    text = re.sub(r"'([^'\\]*(?:\\.[^'\\]*)*)'", replace_single, text)
    text = re.sub(r"([\{,])\s*([A-Za-z_$][\w$\-+]*)\s*:", r'\1 "\2":', text)
    text = re.sub(r"\bNaN\b", "null", text)
    text = re.sub(r"\bInfinity\b", "null", text)
    text = re.sub(r"\bundefined\b", "null", text)
    return text


def find_window_assignment(js_text: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = re.compile(
        r"window(?:\[['\"](?P<qname>[^'\"]+)['\"]\]|\.(?P<dname>[A-Za-z_$][\w$]*))\s*=\s*",
        re.MULTILINE,
    )
    match = pattern.search(js_text)
    if not match:
        return None, None

    variable = match.group("qname") or match.group("dname")
    start = match.end()

    obj_start: Optional[int] = None
    opener: Optional[str] = None

    for i in range(start, len(js_text)):
        if js_text[i] in ("{", "["):
            obj_start = i
            opener = js_text[i]
            break
        if js_text[i] == ";":
            break

    if obj_start is None or opener is None:
        return variable, None

    closer = "}" if opener == "{" else "]"
    depth = 0
    in_str: Optional[str] = None
    escape = False

    for i in range(obj_start, len(js_text)):
        ch = js_text[i]

        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == in_str:
                in_str = None
            continue

        if ch in ("'", '"'):
            in_str = ch
            continue

        if ch == opener:
            depth += 1
        elif ch == closer:
            depth -= 1
            if depth == 0:
                return variable, js_text[obj_start : i + 1]

    return variable, None


def parse_js_file(path: Path) -> Tuple[Optional[str], Optional[Any], Optional[str]]:
    try:
        key = (str(path.resolve()), path.stat().st_mtime)
    except Exception:
        key = (str(path), -1.0)

    cached = _PARSED_JS_CACHE.get(key)
    if cached is not None:
        return cached

    text = read_text(path)
    variable, object_text = find_window_assignment(text)

    if object_text is None:
        result = (variable, None, "No window[...] object assignment found")
        _PARSED_JS_CACHE[key] = result
        return result

    try:
        result = (variable, json.loads(object_text), None)
    except Exception:
        try:
            result = (variable, json.loads(js_like_to_json(object_text)), None)
        except Exception as exc:
            result = (variable, None, f"Could not parse JS object: {exc}")

    _PARSED_JS_CACHE[key] = result
    return result


def get_region_container(parsed: Any, region_level: str = "auto") -> Tuple[Dict[str, Any], str]:
    """
    Return the dictionary whose keys are actual region names.

    Two LUTO dashboard JS layouts are supported:

    1. Older/flat layout:
        {
            "Burdekin": {...},
            "Fitzroy": {...},
            "ACT": {...}
        }

    2. Newer/nested layout:
        {
            "region_NRM": {
                "Burdekin": {...},
                "Fitzroy": {...},
                "ACT": {...}
            },
            "region_state": {
                "Queensland": {...},
                "Victoria": {...}
            }
        }

    If region_level is "auto", the script prefers "region_NRM" when present,
    then "region_state", then falls back to the flat layout.
    """
    if not isinstance(parsed, dict):
        return {}, "none"

    if region_level and region_level.lower() != "auto":
        selected = parsed.get(region_level)
        if isinstance(selected, dict):
            return selected, region_level
        # If the requested region level does not exist, fall back to flat layout.
        return parsed, "flat"

    preferred_levels = ["region_NRM", "region_state", "region_lga", "region_sa2", "region_sa3", "region_sa4"]

    for level in preferred_levels:
        selected = parsed.get(level)
        if isinstance(selected, dict):
            return selected, level

    return parsed, "flat"


def looks_like_year(value: Any) -> bool:
    try:
        year = int(float(value))
        return 1900 <= year <= 2200
    except Exception:
        return False


def to_year(value: Any) -> Optional[int]:
    try:
        return int(float(value))
    except Exception:
        return None


def to_float(value: Any) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def is_series(obj: Any) -> bool:
    """Return True for both old [year, value] series and newer Highcharts {x, y} point series."""
    if not isinstance(obj, dict):
        return False
    data = obj.get("data")
    if not isinstance(data, list):
        return False
    if not data:
        return True

    first = data[0]
    if isinstance(first, list) and len(first) >= 2 and looks_like_year(first[0]):
        return True
    if isinstance(first, dict) and looks_like_year(first.get("x")):
        return True
    return False


def iter_series_points(data: Any):
    """Yield (year, value) from series data stored as [year, value] or {x, y}."""
    if not isinstance(data, list):
        return

    for point in data:
        if isinstance(point, list) and len(point) >= 2:
            year = to_year(point[0])
            value = to_float(point[1])
        elif isinstance(point, dict):
            # Current annualised LUTO economics exports store every year as
            # Highcharts points. Some intermediate years may include an
            # `opacity` field for display styling, but they are still values
            # shown in the dashboard tooltip and must be exported when the
            # requested range includes them. Therefore do NOT skip opacity
            # points.
            year = to_year(point.get("x"))
            value = to_float(point.get("y"))
        else:
            continue

        if year is not None:
            yield year, value


def collect_available_years(obj: Any) -> List[int]:
    years: List[int] = []

    if is_series(obj):
        for year, _value in iter_series_points(obj.get("data", []) or []):
            years.append(year)
        return years

    if isinstance(obj, list):
        for item in obj:
            years.extend(collect_available_years(item))
        return years

    if isinstance(obj, dict):
        for value in obj.values():
            years.extend(collect_available_years(value))
        return years

    return years


def warn_if_requested_years_outside_source(
    run_name: str,
    available_years: List[int],
    start_year: Optional[int],
    end_year: Optional[int],
) -> None:
    clean_years = sorted(set(y for y in available_years if y is not None))

    if not clean_years:
        log(f"WARNING: [{run_name}] No years were detected in the source data.")
        return

    min_available = clean_years[0]
    max_available = clean_years[-1]

    if start_year is not None and start_year < min_available:
        log(
            f"WARNING: [{run_name}] Requested start year is {start_year}, "
            f"but the earliest available source year is {min_available}. "
            f"The output will start from {min_available}."
        )

    if end_year is not None and end_year > max_available:
        log(
            f"WARNING: [{run_name}] Requested end year is {end_year}, "
            f"but the latest available source year is {max_available}. "
            f"The output will end at {max_available}."
        )

    if start_year is not None and end_year is not None:
        in_range = [y for y in clean_years if start_year <= y <= end_year]
        if not in_range:
            log(
                f"WARNING: [{run_name}] Requested range {start_year}-{end_year} "
                f"does not overlap available source years {min_available}-{max_available}."
            )


def walk_series(obj: Any, rows: List[Dict[str, Any]], region: str, context_path: List[str]) -> None:
    if is_series(obj):
        series_name = str(obj.get("name", "value"))
        if context_path:
            series_name = " > ".join(context_path + [series_name])

        for year, value in iter_series_points(obj.get("data", []) or []):
            rows.append({"region": region, "year": year, "series_name": series_name, "value": value})
        return

    if isinstance(obj, list):
        for item in obj:
            walk_series(item, rows, region, context_path)
        return

    if isinstance(obj, dict):
        skip_keys = {"name", "data", "type", "color", "visible", "tooltip", "yAxis"}
        for key, value in obj.items():
            if key in skip_keys:
                continue
            walk_series(value, rows, region, context_path + [str(key)])


def make_wide_table(rows: List[Dict[str, Any]], start_year: Optional[int], end_year: Optional[int]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["Category"])

    df = pd.DataFrame(rows)
    df["year"] = pd.to_numeric(df["year"], errors="coerce")

    if start_year is not None:
        df = df[df["year"] >= start_year]

    if end_year is not None:
        df = df[df["year"] <= end_year]

    if df.empty:
        return pd.DataFrame(columns=["Category"])

    table = df.pivot_table(index="year", columns="series_name", values="value", aggfunc="sum").reset_index()
    table.columns = [str(c) for c in table.columns]
    table = table.rename(columns={"year": "Category"})

    try:
        table["Category"] = pd.to_numeric(table["Category"], errors="coerce").astype("Int64")
    except Exception:
        pass

    observed = [c for c in table.columns if c != "Category"]
    ordered = [s for s in SERIES_ORDER_SUM if s in observed]
    ordered += [c for c in observed if c not in ordered]

    table = table[["Category"] + ordered].sort_values("Category")
    table = table.astype(object).where(pd.notna(table), "")
    return table


def find_file(data_dir: Path, filename: str) -> Optional[Path]:
    direct = data_dir / filename
    if direct.exists():
        return direct
    matches = list(data_dir.rglob(filename))
    return matches[0] if matches else None


def load_overview_tables(
    data_dir: Path,
    region: str,
    start_year: Optional[int],
    end_year: Optional[int],
    region_level: str = "auto",
) -> Dict[str, pd.DataFrame]:
    tables: Dict[str, pd.DataFrame] = {}

    for group, filename in OVERVIEW_FILES.items():
        file_path = find_file(data_dir, filename)

        if file_path is None:
            tables[group] = pd.DataFrame(columns=["Category"])
            continue

        _var, parsed, err = parse_js_file(file_path)
        if parsed is None or not isinstance(parsed, dict):
            log(f"WARNING: Could not parse {file_path}: {err}")
            tables[group] = pd.DataFrame(columns=["Category"])
            continue

        region_container, _actual_level = get_region_container(parsed, region_level)

        if region not in region_container:
            tables[group] = pd.DataFrame(columns=["Category"])
            continue

        rows: List[Dict[str, Any]] = []
        walk_series(region_container[region], rows, region, context_path=[])
        tables[group] = make_wide_table(rows, start_year, end_year)

    return tables


def load_economics_group_profit_table(
    data_dir: Path,
    region: str,
    start_year: Optional[int],
    end_year: Optional[int],
    region_level: str = "auto",
) -> pd.DataFrame:
    """
    Load direct group-level profit values from Economics_Sum.js.

    This is important because LUTO2's dashboard displays these values directly.
    Reconstructing Ag/Ag Mgt/Non-Ag profit from overview components can be close
    but not exactly identical.

    Example:
        Queensland Ag profit 2050 is read from:
            Economics_Sum.js -> region_state -> Queensland -> Agricultural Land-use
    """
    file_path = find_file(data_dir, ECONOMICS_GROUP_PROFIT_FILE)

    if file_path is None:
        return pd.DataFrame(columns=["Category"])

    _var, parsed, err = parse_js_file(file_path)
    if parsed is None or not isinstance(parsed, dict):
        log(f"WARNING: Could not parse {file_path}: {err}")
        return pd.DataFrame(columns=["Category"])

    region_container, _actual_level = get_region_container(parsed, region_level)

    if region not in region_container:
        return pd.DataFrame(columns=["Category"])

    rows: List[Dict[str, Any]] = []
    walk_series(region_container[region], rows, region, context_path=[])
    return make_wide_table(rows, start_year, end_year)


def get_numeric_series_aligned_by_category(
    source_table: pd.DataFrame,
    target_categories: pd.Series,
    column_name: str,
) -> pd.Series:
    """
    Return source_table[column_name] aligned to target_categories using Category as key.
    Missing values remain NaN.
    """
    if (
        source_table is None
        or source_table.empty
        or "Category" not in source_table.columns
        or column_name not in source_table.columns
    ):
        return pd.Series([float("nan")] * len(target_categories), index=target_categories.index)

    src = source_table[["Category", column_name]].copy()
    src["Category"] = pd.to_numeric(src["Category"], errors="coerce")
    src[column_name] = pd.to_numeric(src[column_name], errors="coerce")
    src = src.dropna(subset=["Category"])

    lookup = dict(zip(src["Category"].astype(int), src[column_name]))

    categories = pd.to_numeric(target_categories, errors="coerce")
    values = []
    for cat in categories:
        if pd.isna(cat):
            values.append(float("nan"))
        else:
            values.append(lookup.get(int(cat), float("nan")))

    return pd.Series(values, index=target_categories.index, dtype="float64")


def any_table_has_values(tables: Dict[str, pd.DataFrame]) -> bool:
    """Return True if at least one group table has data beyond a Category header."""
    for table in tables.values():
        if table is not None and not table.empty and "Category" in table.columns:
            if len(table) > 0 and len(table.columns) > 1:
                return True
    return False


def extract_year_totals_by_prefix(
    data_dir: Path,
    filename: str,
    region: str,
    region_level: str,
    start_year: Optional[int],
    end_year: Optional[int],
    include_prefix: str,
) -> pd.Series:
    """
    Sum year values from one JS file for series whose path starts with include_prefix.

    This is used for Ag Mgt renewable revenue/cost breakdown columns. The prefix uses
    the direct dashboard hierarchy, e.g.:

        Onshore Wind > ALL >
        Onshore Wind > ALL > Operating Cost >
        Onshore Wind > ALL > Capital expenditure >

    The use of the ALL level avoids double-counting Dryland/Irrigated subtotals.
    """
    file_path = find_file(data_dir, filename)
    if file_path is None:
        return pd.Series(dtype="float64")

    _var, parsed, err = parse_js_file(file_path)
    if parsed is None or not isinstance(parsed, dict):
        log(f"WARNING: Could not parse {file_path}: {err}")
        return pd.Series(dtype="float64")

    region_container, _actual_level = get_region_container(parsed, region_level)
    if region not in region_container:
        return pd.Series(dtype="float64")

    rows: List[Dict[str, Any]] = []
    walk_series(region_container[region], rows, region, context_path=[])

    if not rows:
        return pd.Series(dtype="float64")

    df = pd.DataFrame(rows)
    if df.empty or "series_name" not in df.columns:
        return pd.Series(dtype="float64")

    df = df[df["series_name"].astype(str).str.startswith(include_prefix)].copy()
    if df.empty:
        return pd.Series(dtype="float64")

    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["year", "value"])

    if start_year is not None:
        df = df[df["year"] >= start_year]
    if end_year is not None:
        df = df[df["year"] <= end_year]

    if df.empty:
        return pd.Series(dtype="float64")

    totals = df.groupby("year")["value"].sum(min_count=1)
    totals.index = totals.index.astype(int)
    return totals.sort_index()


def extract_year_totals_exact_series(
    data_dir: Path,
    filename: str,
    region: str,
    region_level: str,
    start_year: Optional[int],
    end_year: Optional[int],
    exact_series_name: str,
) -> pd.Series:
    """
    Return year values from one JS file for an exact flattened series name.

    This is used for values where the dashboard source already provides a clean
    ALL total, for example:

        Economics_overview_Non_Ag.js -> Cost > ALL
    """
    file_path = find_file(data_dir, filename)
    if file_path is None:
        return pd.Series(dtype="float64")

    _var, parsed, err = parse_js_file(file_path)
    if parsed is None or not isinstance(parsed, dict):
        log(f"WARNING: Could not parse {file_path}: {err}")
        return pd.Series(dtype="float64")

    region_container, _actual_level = get_region_container(parsed, region_level)
    if region not in region_container:
        return pd.Series(dtype="float64")

    rows: List[Dict[str, Any]] = []
    walk_series(region_container[region], rows, region, context_path=[])

    if not rows:
        return pd.Series(dtype="float64")

    df = pd.DataFrame(rows)
    if df.empty or "series_name" not in df.columns:
        return pd.Series(dtype="float64")

    df = df[df["series_name"].astype(str) == exact_series_name].copy()
    if df.empty:
        return pd.Series(dtype="float64")

    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["year", "value"])

    if start_year is not None:
        df = df[df["year"] >= start_year]
    if end_year is not None:
        df = df[df["year"] <= end_year]

    if df.empty:
        return pd.Series(dtype="float64")

    totals = df.groupby("year")["value"].sum(min_count=1)
    totals.index = totals.index.astype(int)
    return totals.sort_index()


def align_year_series_to_table(series: pd.Series, table: pd.DataFrame) -> pd.Series:
    """Align a year-indexed Series to table['Category'] without inventing missing values."""
    if table is None or table.empty or "Category" not in table.columns:
        return pd.Series(dtype="float64")

    values: List[Any] = []
    for category in pd.to_numeric(table["Category"], errors="coerce"):
        if pd.isna(category):
            values.append(float("nan"))
        else:
            values.append(series.get(int(category), float("nan")))
    return pd.Series(values, index=table.index, dtype="float64")


def apply_dashboard_source_overrides(
    tables: Dict[str, pd.DataFrame],
    data_dir: Path,
    region: str,
    region_level: str,
    start_year: Optional[int],
    end_year: Optional[int],
) -> Dict[str, pd.DataFrame]:
    """
    Apply the reviewed dashboard source mappings used by the reporting workbook.

    Corrections:

    1. Sum -> Non-Agricultural Land-use (cost)
       Use Economics_overview_Non_Ag.js -> Cost > ALL.

    2. Sum -> Agricultural Management (cost)
       Use Economics_Am_cost.js -> ALL > ALL > ALL > <land-use> and sum only
       those leaf series. This is the total shown by LUTO2 when Category is
       Ag Mgt, Map Type is Cost, and Ag Mgt / Water / Cost selections are ALL.
       It must not be added to Operating Cost and Capital expenditure again,
       because the ALL total already contains those components.

    3. Ag Mgt -> Onshore Wind / Utility Solar PV
       Use direct profit totals from Economics_Am_profit.js and label the two
       columns under a Profit group header.

    The original dashboard Profit column in the Sum table is retained as a
    direct extracted value. It is not recalculated from the corrected component
    columns, because each dashboard chart is treated as the authoritative source
    for its own displayed value.
    """
    out = dict(tables)

    # --- Sum table visible cost overrides ---
    sum_table = out.get("Sum")
    if sum_table is not None and not sum_table.empty and "Category" in sum_table.columns:
        local = sum_table.copy()

        # Non-Ag cost shown in the visible Non-Ag / Cost / ALL chart.
        corrected_nonag_cost = extract_year_totals_exact_series(
            data_dir=data_dir,
            filename="Economics_overview_Non_Ag.js",
            region=region,
            region_level=region_level,
            start_year=start_year,
            end_year=end_year,
            exact_series_name="Cost > ALL",
        )
        aligned_nonag_cost = align_year_series_to_table(corrected_nonag_cost, local)
        if not aligned_nonag_cost.empty and aligned_nonag_cost.notna().any():
            if "Non-Agricultural Land-use (cost)" not in local.columns:
                local["Non-Agricultural Land-use (cost)"] = ""
            existing = pd.to_numeric(local["Non-Agricultural Land-use (cost)"], errors="coerce")
            local["Non-Agricultural Land-use (cost)"] = aligned_nonag_cost.where(
                aligned_nonag_cost.notna(), existing
            )

        # Ag Mgt cost shown in the visible Ag Mgt / Cost / ALL chart.
        # The selected ALL path contains the complete total, so only the leaf
        # series beneath ALL > ALL > ALL are summed. Opex/capex must not be added
        # again or the result is doubled.
        corrected_am_cost = extract_year_totals_by_prefix(
            data_dir=data_dir,
            filename="Economics_Am_cost.js",
            region=region,
            region_level=region_level,
            start_year=start_year,
            end_year=end_year,
            include_prefix="ALL > ALL > ALL >",
        )
        aligned_am_cost = align_year_series_to_table(corrected_am_cost, local)
        if not aligned_am_cost.empty and aligned_am_cost.notna().any():
            if "Agricultural Management (cost)" not in local.columns:
                local["Agricultural Management (cost)"] = ""
            existing = pd.to_numeric(local["Agricultural Management (cost)"], errors="coerce")
            local["Agricultural Management (cost)"] = aligned_am_cost.where(
                aligned_am_cost.notna(), existing
            )

        local = local.astype(object).where(pd.notna(local), "")
        out["Sum"] = local

    # --- Ag Mgt wind/solar direct-profit overrides ---
    ag_mgt_table = out.get("Ag Mgt")
    if ag_mgt_table is not None and not ag_mgt_table.empty and "Category" in ag_mgt_table.columns:
        local = ag_mgt_table.copy()
        profit_columns: Dict[str, pd.Series] = {}

        for tech in AG_MGT_BREAKDOWN_TECHS:
            direct_profit = extract_year_totals_by_prefix(
                data_dir=data_dir,
                filename="Economics_Am_profit.js",
                region=region,
                region_level=region_level,
                start_year=start_year,
                end_year=end_year,
                include_prefix=f"{tech} > ALL >",
            )
            aligned_profit = align_year_series_to_table(direct_profit, local)
            if not aligned_profit.empty and aligned_profit.notna().any():
                existing = (
                    pd.to_numeric(local[tech], errors="coerce")
                    if tech in local.columns
                    else pd.Series([float("nan")] * len(local), index=local.index)
                )
                profit_columns[tech] = aligned_profit.where(aligned_profit.notna(), existing)

        # Preserve the original column positions while making the meaning clear
        # in Excel: columns H/I appear under a merged Profit heading.
        for tech, values in profit_columns.items():
            source_col = tech
            target_col = f"Profit||{tech}"
            if source_col in local.columns:
                insert_at = local.columns.get_loc(source_col)
                local[source_col] = values
                local = local.rename(columns={source_col: target_col})
                # rename preserves the position; insert_at retained for clarity
                _ = insert_at
            else:
                local[target_col] = values

        local = local.astype(object).where(pd.notna(local), "")
        out["Ag Mgt"] = local

    return out

def add_ag_mgt_revenue_cost_breakdown(
    ag_mgt_table: pd.DataFrame,
    data_dir: Path,
    region: str,
    region_level: str,
    start_year: Optional[int],
    end_year: Optional[int],
) -> pd.DataFrame:
    """
    Add the extra Ag Mgt revenue/cost breakdown columns requested in the template.

    The original Ag Mgt overview table shows net/overview values. This function appends
    renewable revenue and cost components for Onshore Wind and Utility Solar PV:

        Revenue
        Costs (opex)
        Costs (capex)
        Costs (other)

    Costs (other) is included as a blank placeholder because the current JS source
    contains Operating Cost and Capital expenditure, but not an "other" cost category.
    """
    if ag_mgt_table is None or ag_mgt_table.empty or "Category" not in ag_mgt_table.columns:
        base_table = pd.DataFrame(columns=["Category"])
    else:
        base_table = ag_mgt_table.copy()

    categories = pd.to_numeric(base_table["Category"], errors="coerce") if "Category" in base_table.columns else pd.Series(dtype="float64")
    year_index = []
    for value in categories:
        if pd.notna(value):
            year_index.append(int(value))

    # If the main table has no year rows, return the original table with just headers.
    out = base_table.copy()

    def align_series(series: pd.Series) -> List[Any]:
        vals = []
        for y in year_index:
            vals.append(series.get(y, ""))
        return vals

    if "Category" not in out.columns:
        out["Category"] = year_index

    # Ensure output rows exist before adding component columns.
    if len(out) == 0 and year_index:
        out = pd.DataFrame({"Category": year_index})

    for parent, tech in AG_MGT_DETAIL_COLUMNS:
        col_name = f"{parent}||{tech}"

        if parent == "Revenue":
            prefix = f"{tech} > ALL >"
            series = extract_year_totals_by_prefix(
                data_dir, "Economics_Am_revenue.js", region, region_level,
                start_year, end_year, prefix
            )
            out[col_name] = align_series(series)

        elif parent == "Costs (opex) ":
            prefix = f"{tech} > ALL > Operating Cost >"
            series = extract_year_totals_by_prefix(
                data_dir, "Economics_Am_cost.js", region, region_level,
                start_year, end_year, prefix
            )
            out[col_name] = align_series(series)

        elif parent == "Costs (capex) ":
            prefix = f"{tech} > ALL > Capital expenditure >"
            series = extract_year_totals_by_prefix(
                data_dir, "Economics_Am_cost.js", region, region_level,
                start_year, end_year, prefix
            )
            out[col_name] = align_series(series)

        else:
            # Current source has no third "other costs" category.
            out[col_name] = [""] * len(out)

    return out


def get_available_years_from_overview_files(data_dir: Path) -> List[int]:
    years: List[int] = []

    for filename in OVERVIEW_FILES.values():
        file_path = find_file(data_dir, filename)
        if file_path is None:
            continue
        _var, parsed, _err = parse_js_file(file_path)
        if parsed is not None:
            years.extend(collect_available_years(parsed))

    return years


def numeric_column_or_nan(table: pd.DataFrame, column_name: str) -> pd.Series:
    """Return a numeric column if present; otherwise return NaNs aligned to the table."""
    if table is None or table.empty:
        return pd.Series(dtype="float64")

    if column_name not in table.columns:
        return pd.Series([float("nan")] * len(table), index=table.index)

    return pd.to_numeric(table[column_name], errors="coerce")


def sum_components_preserve_missing(components: List[pd.Series], length: int, index) -> pd.Series:
    """
    Sum components by row while preserving missingness.

    If all components are missing for a row, the result remains NaN.
    If at least one component exists, missing components are treated as zero
    for that row's sum.
    """
    if not components:
        return pd.Series([float("nan")] * length, index=index)

    frame = pd.concat(components, axis=1)
    all_missing = frame.isna().all(axis=1)
    total = frame.fillna(0.0).sum(axis=1)
    total[all_missing] = float("nan")
    return total


def make_profitability_table_from_sum_components(
    sum_table: pd.DataFrame,
    group_name: str,
    group_profit_table: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    Compute profitability for Sum, Ag, Ag Mgt, and Non-Ag from the clean Sum overview table.

    This avoids the double-counting problem in the detailed source files:
      - Economics_Ag_revenue.js and Economics_Ag_cost.js contain overlapping totals/subtotals.
      - Summing every series in those files overcounts revenue/cost.
      - Therefore profitability should be derived from the already aggregated dashboard Sum components.

    Definitions:
      Sum:
        revenue = Ag revenue + Ag Mgt revenue + Non-Ag revenue
        cost    = Ag cost + Ag Mgt cost + Non-Ag cost + Ag2Ag transition + Ag2Non-Ag transition
        profit  = dashboard Profit column, if available

      Ag:
        revenue = Agricultural Land-use revenue
        cost    = Agricultural Land-use cost + Transition cost (Ag2Ag)

      Ag Mgt:
        revenue = Agricultural Management revenue
        cost    = Agricultural Management cost

      Non-Ag:
        revenue = Non-Agricultural Land-use revenue
        cost    = Non-Agricultural Land-use cost + Transition cost (Ag2Non-Ag)

      Profitability (%) = Profit / Total revenue * 100

    Missing-data rule:
      If all source components for a group/year are missing, the profitability row is left blank
      rather than being forced to zero.
    """
    headers = ["Category", "Total revenue", "Total cost", "Profit", "Profitability (%)"]

    if sum_table is None or sum_table.empty or "Category" not in sum_table.columns:
        return pd.DataFrame(columns=headers)

    table = sum_table.copy()
    out = pd.DataFrame()
    out["Category"] = table["Category"]

    ag_rev = numeric_column_or_nan(table, "Agricultural Land-use (revenue)")
    am_rev = numeric_column_or_nan(table, "Agricultural Management (revenue)")
    nonag_rev = numeric_column_or_nan(table, "Non-Agricultural Land-use (revenue)")

    ag_cost = numeric_column_or_nan(table, "Agricultural Land-use (cost)")
    am_cost = numeric_column_or_nan(table, "Agricultural Management (cost)")
    nonag_cost = numeric_column_or_nan(table, "Non-Agricultural Land-use (cost)")
    transition_ag2ag = numeric_column_or_nan(table, "Transition cost (Ag2Ag)")
    transition_ag2nonag = numeric_column_or_nan(table, "Transition cost (Ag2Non-Ag)")
    dashboard_profit = numeric_column_or_nan(table, "Profit")

    direct_group_profit = pd.Series([float("nan")] * len(table), index=table.index)
    direct_profit_column = GROUP_PROFIT_COLUMN.get(group_name)
    if direct_profit_column:
        direct_group_profit = get_numeric_series_aligned_by_category(
            group_profit_table if group_profit_table is not None else pd.DataFrame(),
            table["Category"],
            direct_profit_column,
        )

    if group_name == "Sum":
        revenue_components = [ag_rev, am_rev, nonag_rev]
        cost_components = [ag_cost, am_cost, nonag_cost, transition_ag2ag, transition_ag2nonag]
        total_revenue = sum_components_preserve_missing(revenue_components, len(table), table.index)
        total_cost = sum_components_preserve_missing(cost_components, len(table), table.index)

        # Use one consistent accounting basis in the derived profitability table.
        # The direct Economics_overview_sum.js Profit series is retained in the
        # main Sum block because it is the value shown by that dashboard graph.
        # However, that direct series still embeds the duplicated Ag Mgt and
        # Non-Ag costs found in the overview source. Mixing it with the corrected
        # detail-chart costs makes Profit != Total revenue + Total cost at the
        # five-year graph points. Therefore this derived table always uses:
        #
        #     Profit = Total revenue + Total cost
        #
        # where costs are negative.
        profit = sum_components_preserve_missing([total_revenue, total_cost], len(table), table.index)

    elif group_name == "Ag":
        group_components = [ag_rev, ag_cost, transition_ag2ag]
        total_revenue = sum_components_preserve_missing([ag_rev], len(table), table.index)
        total_cost = sum_components_preserve_missing([ag_cost, transition_ag2ag], len(table), table.index)
        fallback_profit = sum_components_preserve_missing(group_components, len(table), table.index)
        profit = direct_group_profit.where(direct_group_profit.notna(), fallback_profit)

    elif group_name == "Ag Mgt":
        group_components = [am_rev, am_cost]
        total_revenue = sum_components_preserve_missing([am_rev], len(table), table.index)
        total_cost = sum_components_preserve_missing([am_cost], len(table), table.index)
        fallback_profit = sum_components_preserve_missing(group_components, len(table), table.index)
        profit = direct_group_profit.where(direct_group_profit.notna(), fallback_profit)

    elif group_name == "Non-Ag":
        group_components = [nonag_rev, nonag_cost, transition_ag2nonag]
        total_revenue = sum_components_preserve_missing([nonag_rev], len(table), table.index)
        total_cost = sum_components_preserve_missing([nonag_cost, transition_ag2nonag], len(table), table.index)
        fallback_profit = sum_components_preserve_missing(group_components, len(table), table.index)
        profit = direct_group_profit.where(direct_group_profit.notna(), fallback_profit)

    else:
        return pd.DataFrame(columns=headers)

    # Some current LUTO DATA_REPORT exports do not provide a direct revenue
    # series for every group/region. When direct revenue is missing but both
    # source profit and total-cost components are available, infer revenue using
    # the accounting identity used by the dashboard values:
    #
    #     Profit = Total revenue + Total cost
    #
    # Costs are stored as negative values in the source, so:
    #
    #     Total revenue = Profit - Total cost
    #
    # Use this only for groups where the relationship is stable in the source
    # layout. In particular, Ag Mgt direct profit can differ from the overview
    # revenue/cost pair, so it is deliberately not inferred here.
    if group_name == "Ag":
        # Ag normally has a direct revenue series. Retain the historical fallback
        # only for genuinely missing rows.
        inferred_revenue = profit - total_cost
        inferred_revenue = inferred_revenue.mask(inferred_revenue.abs() < 1.0, 0.0)
        can_infer_revenue = profit.notna() & total_cost.notna()
        total_revenue = total_revenue.where(total_revenue.notna(), inferred_revenue.where(can_infer_revenue))

    elif group_name == "Non-Ag":
        # Queensland has no direct Non-Ag revenue series in the uploaded source.
        # Direct Non-Ag profit is effectively equal to corrected cost + Ag2Non-Ag
        # transition cost, with only small floating-point/source-rounding residuals.
        # Those residuals are not economic revenue. Set revenue to zero when the
        # identity holds within a scale-aware tolerance; otherwise leave it blank
        # rather than inventing a value. Profitability is consequently blank for
        # zero-revenue rows.
        residual = profit - total_cost
        scale = pd.concat([profit.abs(), total_cost.abs()], axis=1).max(axis=1).fillna(0.0)
        tolerance = (scale * 2.0e-6).clip(lower=1.0)
        near_zero_revenue = residual.abs() <= tolerance
        zero_revenue = pd.Series(0.0, index=table.index)
        total_revenue = total_revenue.where(
            total_revenue.notna(),
            zero_revenue.where(near_zero_revenue),
        )

    profitability = (profit / total_revenue) * 100.0
    profitability = profitability.where(total_revenue != 0)

    out["Total revenue"] = total_revenue
    out["Total cost"] = total_cost
    out["Profit"] = profit
    out["Profitability (%)"] = profitability

    # Replace Inf with NA without triggering pandas downcasting FutureWarning.
    for col in ["Total revenue", "Total cost", "Profit", "Profitability (%)"]:
        numeric_col = pd.to_numeric(out[col], errors="coerce")
        numeric_col = numeric_col.mask(numeric_col.isin([float("inf"), float("-inf")]))
        out[col] = numeric_col

    # Replace NaN with blanks for Excel readability.
    out = out.astype(object).where(pd.notna(out), "")

    # If the table has no meaningful revenue/cost/profit values, return empty.
    value_cols = ["Total revenue", "Total cost", "Profit", "Profitability (%)"]
    numeric_any = False
    for col in value_cols:
        numeric_values = pd.to_numeric(out[col], errors="coerce")
        if numeric_values.notna().any():
            numeric_any = True
            break

    if not numeric_any:
        return pd.DataFrame(columns=headers)

    return out[headers]


def sum_tables(tables: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Sum regional overview tables while preserving blanks.

    If all regional values are missing for a column/year, the state-total value
    remains blank instead of becoming zero.
    """
    valid = [t.copy() for t in tables if t is not None and not t.empty and "Category" in t.columns]

    if not valid:
        return pd.DataFrame(columns=["Category"])

    normalised: List[pd.DataFrame] = []

    for table in valid:
        local = table.copy()
        local["Category"] = pd.to_numeric(local["Category"], errors="coerce")
        local = local.dropna(subset=["Category"])

        for col in local.columns:
            if col == "Category":
                continue
            local[col] = pd.to_numeric(local[col], errors="coerce")

        normalised.append(local)

    if not normalised:
        return pd.DataFrame(columns=["Category"])

    combined = pd.concat(normalised, ignore_index=True, sort=False)
    value_cols = [c for c in combined.columns if c != "Category"]

    summed = combined.groupby("Category", as_index=False)[value_cols].sum(min_count=1)
    summed = summed.sort_values("Category")

    try:
        summed["Category"] = summed["Category"].astype(int)
    except Exception:
        pass

    summed = summed.astype(object).where(pd.notna(summed), "")
    return summed


def write_block(ws, start_row: int, title: str, table: pd.DataFrame) -> int:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Preserve the intended header layout for empty tables.
    # Earlier versions replaced every empty DataFrame with only ["Category"],
    # which made empty profitability blocks lose their required columns
    # (Total revenue, Total cost, Profit, Profitability (%)).
    if table is None:
        table = pd.DataFrame(columns=["Category"])
    elif table.empty and len(table.columns) == 0:
        table = pd.DataFrame(columns=["Category"])

    ncols = max(1, len(table.columns))
    has_grouped_headers = any("||" in str(c) for c in table.columns)

    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).fill = title_fill
    ws.cell(start_row, 1).font = white_font
    ws.cell(start_row, 1).alignment = Alignment(horizontal="left")

    if has_grouped_headers:
        # Template-style two-row header:
        #   row start_row: section title plus group labels such as Revenue / Costs
        #   row start_row+1: Category plus child labels such as Onshore Wind / Utility Solar PV
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(start_row, col_idx)
            cell.fill = title_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        parents = []
        children = []
        for col in table.columns:
            col_s = str(col)
            if "||" in col_s:
                parent, child = col_s.split("||", 1)
                parents.append(parent)
                children.append(child)
            else:
                parents.append("")
                children.append(col_s)

        # Write and merge consecutive parent group labels.
        col = 1
        while col <= ncols:
            parent = parents[col - 1]
            if parent:
                end_col = col
                while end_col + 1 <= ncols and parents[end_col] == parent:
                    end_col += 1
                ws.cell(start_row, col, parent)
                if end_col > col:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=end_col)
                col = end_col + 1
            else:
                col += 1

        header_row = start_row + 1
        for col_idx, child_name in enumerate(children, start=1):
            cell = ws.cell(header_row, col_idx, child_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        data_start = start_row + 2

    else:
        if ncols > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)

        header_row = start_row + 1
        for col_idx, col_name in enumerate(table.columns, start=1):
            cell = ws.cell(header_row, col_idx, col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        data_start = start_row + 2

    for r_idx, row in enumerate(table.itertuples(index=False, name=None), start=data_start):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, None if value == "" else value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if c_idx == 1:
                cell.font = Font(bold=True)
            elif isinstance(cell.value, float):
                cell.number_format = "0.###############"

    return data_start + len(table) + 2


def format_sheet(ws) -> None:
    ws.freeze_panes = "A3"

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row_idx, col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)


def write_region_sheet(ws, region: str, tables: Dict[str, pd.DataFrame], state: str, run_name: str) -> None:
    ws.cell(1, 1, f"{run_name} | {state} | {region}")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="left")

    row = 3
    for group in GROUP_ORDER:
        row = write_block(ws, row, group, tables.get(group, pd.DataFrame(columns=["Category"])))

        profitability_table = tables.get(
            f"{group} - Profitability",
            pd.DataFrame(columns=["Category", "Total revenue", "Total cost", "Profit", "Profitability (%)"]),
        )
        # Always write the profitability block so each sheet visibly contains
        # Sum/Ag/Ag Mgt/Non-Ag profitability, even when a region has no values
        # for that group/year.
        row = write_block(ws, row, f"{group} - Profitability", profitability_table)

    format_sheet(ws)


def load_state_mapping(csv_path: Optional[Path]) -> Dict[str, str]:
    mapping = dict(REGION_TO_STATE)

    if csv_path is None:
        return mapping

    if not csv_path.exists():
        raise FileNotFoundError(f"State mapping CSV not found: {csv_path}")

    df = pd.read_csv(csv_path)
    cols = {c.lower(): c for c in df.columns}

    if "region" not in cols or "state" not in cols:
        raise ValueError("State mapping CSV must contain columns: region,state")

    for _, row in df.iterrows():
        region = str(row[cols["region"]]).strip()
        state = str(row[cols["state"]]).strip()
        if region and state:
            mapping[region] = state

    return mapping


def get_regions_from_overview_sum(data_dir: Path, include_national: bool, region_level: str = "auto") -> List[str]:
    file_path = find_file(data_dir, "Economics_overview_sum.js")
    if file_path is None:
        raise FileNotFoundError(f"Could not find Economics_overview_sum.js under {data_dir}")

    _var, parsed, err = parse_js_file(file_path)
    if parsed is None or not isinstance(parsed, dict):
        raise ValueError(f"Could not parse {file_path}: {err}")

    region_container, actual_level = get_region_container(parsed, region_level)
    regions = sorted(str(r) for r in region_container.keys())

    if not include_national:
        regions = [r for r in regions if r != "AUSTRALIA"]

    log(f"Region level used: {actual_level}")
    log(f"Regions detected: {len(regions)}")
    return regions


def find_data_dir_inside_run(run_dir: Path) -> Optional[Path]:
    candidates = [
        run_dir / "DATA_REPORT" / "data",
        run_dir / "data",
    ]

    for candidate in candidates:
        if (candidate / "Economics_overview_sum.js").exists():
            return candidate

    matches = list(run_dir.rglob("Economics_overview_sum.js"))
    if matches:
        return matches[0].parent

    return None


def discover_runs(reports_base_dir: Path, run_prefix: str, run_names: List[str]) -> List[Tuple[str, Path]]:
    if not reports_base_dir.exists():
        raise FileNotFoundError(f"Reports base directory does not exist: {reports_base_dir}")

    wanted = {r.lower() for r in run_names}
    candidates = [p for p in reports_base_dir.iterdir() if p.is_dir() and p.name.startswith(run_prefix)]

    if not candidates:
        candidates = [p for p in reports_base_dir.rglob(f"{run_prefix}*") if p.is_dir()]

    runs = []
    seen = set()

    for run_dir in sorted(candidates):
        if run_dir in seen:
            continue
        seen.add(run_dir)

        if wanted and run_dir.name.lower() not in wanted:
            continue

        data_dir = find_data_dir_inside_run(run_dir)
        if data_dir is None:
            log(f"WARNING: Skipping {run_dir.name}; no Economics_overview_sum.js found.")
            continue

        runs.append((run_dir.name, data_dir))

    return runs


def infer_run_name_from_data_dir(data_dir: Path, fallback: str = "Run") -> str:
    parts = list(data_dir.parts)
    lower = [p.lower() for p in parts]

    if "data_report" in lower:
        idx = lower.index("data_report")
        if idx > 0:
            return sanitize_filename(parts[idx - 1])

    return sanitize_filename(fallback)


def build_run_list(args: argparse.Namespace) -> List[Tuple[str, Path]]:
    if args.reports_base_dir:
        return discover_runs(Path(args.reports_base_dir), args.run_prefix, args.run_names or [])

    if args.data_dir:
        data_dir = Path(args.data_dir)

        if data_dir.name.lower() == "map_layers" and data_dir.parent.name.lower() == "data":
            log("WARNING: You supplied DATA_REPORT/data/map_layers. Switching to parent DATA_REPORT/data.")
            data_dir = data_dir.parent

        if not data_dir.exists():
            raise FileNotFoundError(f"Data directory does not exist: {data_dir}")

        run_name = args.output_prefix or infer_run_name_from_data_dir(data_dir, fallback="Run")
        return [(sanitize_filename(run_name), data_dir)]

    raise ValueError("You must provide either --reports-base-dir or --data-dir.")


def process_one_run(
    run_name: str,
    data_dir: Path,
    output_dir: Path,
    mapping: Dict[str, str],
    wanted_states: set[str],
    wanted_regions: set[str],
    start_year: Optional[int],
    end_year: Optional[int],
    include_national: bool,
    region_level: str = "auto",
) -> List[Path]:
    log("")
    log("=" * 80)
    log(f"Processing {run_name}")
    log(f"Data folder: {data_dir}")
    log("=" * 80)

    available_years = get_available_years_from_overview_files(data_dir)
    warn_if_requested_years_outside_source(run_name, available_years, start_year, end_year)

    regions = get_regions_from_overview_sum(data_dir, include_national=include_national, region_level=region_level)

    by_state: Dict[str, List[str]] = {}
    for region in regions:
        if wanted_regions and region.lower() not in wanted_regions:
            continue

        state = mapping.get(region, "Unknown_State")

        if wanted_states and state.lower() not in wanted_states:
            continue

        by_state.setdefault(state, []).append(region)

    outputs: List[Path] = []

    for state, state_regions in sorted(by_state.items()):
        state_folder = output_dir / sanitize_filename(state)
        state_folder.mkdir(parents=True, exist_ok=True)

        output_file = state_folder / f"{sanitize_filename(run_name)}_{sanitize_filename(state)}_Economics_Tab_Overview.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # Reserve the state-total sheet name only when a state-total sheet will actually be created.
        # For single-region states where region name == state name, e.g. Northern Territory,
        # reserving the state name causes an unnecessary sheet name like "Northern Territory_1".
        used_sheets: set[str] = {sanitize_sheet_name(state).lower()} if len(state_regions) > 1 else set()
        state_tables_by_group: Dict[str, List[pd.DataFrame]] = {g: [] for g in GROUP_ORDER}

        for region in sorted(state_regions):
            # For multi-region states, region sheets should use region_NRM.
            # For single-region states (e.g. ACT, Northern Territory), LUTO2 also
            # provides direct state-level values under region_state. Use those direct
            # state values when available so the workbook matches LUTO2 directly.
            source_region = region
            source_region_level = region_level

            if len(state_regions) == 1:
                direct_state_tables = load_overview_tables(
                    data_dir,
                    state,
                    start_year,
                    end_year,
                    region_level="region_state",
                )
                if any_table_has_values(direct_state_tables):
                    tables = direct_state_tables
                    source_region = state
                    source_region_level = "region_state"
                else:
                    tables = load_overview_tables(data_dir, region, start_year, end_year, region_level=region_level)
            else:
                tables = load_overview_tables(data_dir, region, start_year, end_year, region_level=region_level)

            # Apply dashboard-aligned source mappings before derived calculations.
            tables = apply_dashboard_source_overrides(
                tables,
                data_dir,
                source_region,
                source_region_level,
                start_year,
                end_year,
            )

            # Add the Ag Mgt revenue/cost breakdown requested in the template.
            tables["Ag Mgt"] = add_ag_mgt_revenue_cost_breakdown(
                tables.get("Ag Mgt", pd.DataFrame(columns=["Category"])),
                data_dir,
                source_region,
                source_region_level,
                start_year,
                end_year,
            )

            # Profitability uses dashboard-level Sum components for revenue/cost, but uses
            # direct group-level profit from Economics_Sum.js where available.
            # This matches the values displayed directly in LUTO2.
            sum_table = tables.get("Sum", pd.DataFrame(columns=["Category"]))
            group_profit_table = load_economics_group_profit_table(
                data_dir, source_region, start_year, end_year, region_level=source_region_level
            )
            for group in GROUP_ORDER:
                tables[f"{group} - Profitability"] = make_profitability_table_from_sum_components(
                    sum_table,
                    group,
                    group_profit_table=group_profit_table,
                )

            for group, table in tables.items():
                if group in state_tables_by_group and table is not None and not table.empty:
                    state_tables_by_group[group].append(table)

            sheet_name = unique_sheet_name(region, used_sheets)
            ws = wb.create_sheet(sheet_name)
            write_region_sheet(ws, region, tables, state, run_name)

        if len(state_regions) > 1:
            # Prefer LUTO2's own state-level dashboard values when available.
            # This avoids small discrepancies caused by summing NRM sheets and matches
            # what users see when selecting the state directly in LUTO2.
            direct_state_tables = load_overview_tables(
                data_dir,
                state,
                start_year,
                end_year,
                region_level="region_state",
            )

            if any_table_has_values(direct_state_tables):
                total_tables = direct_state_tables
                total_source_note = "direct region_state values"
            else:
                total_tables = {
                    group: sum_tables(group_tables)
                    for group, group_tables in state_tables_by_group.items()
                }
                total_source_note = "summed region sheets"

            # Apply the same dashboard-aligned mappings to the state totals.
            total_tables = apply_dashboard_source_overrides(
                total_tables,
                data_dir,
                state,
                "region_state",
                start_year,
                end_year,
            )

            total_tables["Ag Mgt"] = add_ag_mgt_revenue_cost_breakdown(
                total_tables.get("Ag Mgt", pd.DataFrame(columns=["Category"])),
                data_dir,
                state,
                "region_state",
                start_year,
                end_year,
            )

            state_sum_table = total_tables.get("Sum", pd.DataFrame(columns=["Category"]))
            state_group_profit_table = load_economics_group_profit_table(
                data_dir,
                state,
                start_year,
                end_year,
                region_level="region_state",
            )

            for group in GROUP_ORDER:
                total_tables[f"{group} - Profitability"] = make_profitability_table_from_sum_components(
                    state_sum_table,
                    group,
                    group_profit_table=state_group_profit_table,
                )

            ws_total = wb.create_sheet(sanitize_sheet_name(state))
            write_region_sheet(ws_total, state, total_tables, state, run_name)
            total_note = f" + 1 state-total sheet ({total_source_note})"
        else:
            total_note = ""

        wb.save(output_file)
        outputs.append(output_file)
        log(f"Saved: {output_file} ({len(state_regions)} region sheet(s){total_note})")

    return outputs


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Export LUTO2 Economics tab data into state-organised Excel workbooks with Sum, Ag, Ag Mgt, Non-Ag, profitability, and template-level revenue/cost breakdowns."
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument("--reports-base-dir", help="Folder containing Run_* folders.")
    input_group.add_argument("--data-dir", help="Single run DATA_REPORT/data folder.")

    parser.add_argument("--output-dir", required=True, help="Folder where outputs will be saved.")
    parser.add_argument("--output-prefix", default="", help="Optional run/output prefix for single-run mode.")
    parser.add_argument("--run-prefix", default="Run_", help="Prefix used to discover run folders. Default: Run_")
    parser.add_argument("--run-names", nargs="*", default=[], help="Optional list of run folders to process.")
    parser.add_argument("--states", nargs="*", default=[], help="Optional state filter, e.g. Queensland Victoria")
    parser.add_argument("--regions", nargs="*", default=[], help="Optional region filter, e.g. Fitzroy")
    parser.add_argument("--start-year", type=int, default=2020)
    parser.add_argument("--end-year", type=int, default=2050)
    parser.add_argument("--state-map-csv", default="", help="Optional CSV with columns region,state.")
    parser.add_argument("--include-national", action="store_true", help="Include AUSTRALIA in a National workbook.")
    parser.add_argument(
        "--region-level",
        default="auto",
        help=(
            "Region level inside the LUTO JS files. Use auto unless needed. "
            "Typical values: auto, region_NRM, region_state. Default: auto."
        ),
    )

    args = parser.parse_args(argv)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    mapping = load_state_mapping(Path(args.state_map_csv) if args.state_map_csv else None)
    wanted_states = {s.lower() for s in args.states}
    wanted_regions = {r.lower() for r in args.regions}

    try:
        runs = build_run_list(args)
    except Exception as exc:
        log(f"ERROR: {exc}")
        return 1

    if not runs:
        log("ERROR: No run folders/data folders were found.")
        return 1

    log(f"Found {len(runs)} run(s).")

    all_outputs: List[Path] = []
    failures = 0

    for run_name, data_dir in runs:
        try:
            outputs = process_one_run(
                run_name=run_name,
                data_dir=data_dir,
                output_dir=output_dir,
                mapping=mapping,
                wanted_states=wanted_states,
                wanted_regions=wanted_regions,
                start_year=args.start_year,
                end_year=args.end_year,
                include_national=args.include_national,
                region_level=args.region_level,
            )
            all_outputs.extend(outputs)
        except Exception as exc:
            failures += 1
            log(f"ERROR: Failed to process {run_name}: {exc}")

    if all_outputs:
        manifest_path = output_dir / "Economics_Tab_Overview_State_Export_Manifest.csv"
        manifest = pd.DataFrame([{"output_file": str(p), "filename": p.name} for p in all_outputs])
        manifest.to_csv(manifest_path, index=False, encoding="utf-8-sig")
        log("")
        log(f"Manifest saved: {manifest_path}")

    if failures:
        log(f"Finished with {failures} failed run(s).")
        return 10

    log("Finished successfully.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
